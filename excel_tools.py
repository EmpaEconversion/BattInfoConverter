"""
excel_tools.py
A drop-in replacement for ``pandas.read_excel`` that

1. preserves the number of decimals users see in Excel,
2. mimics pandas’ header handling (Unnamed columns + de-duplication).

Import like::

    from excel_tools import read_excel_preserve_decimals as read_excel
"""

from decimal import Decimal, ROUND_HALF_UP
from typing import Any, List, Sequence

import pandas as pd
from openpyxl import load_workbook


try:  # official path (openpyxl ≥ 3.1)
    from openpyxl.utils.formatting import format_cell  # type: ignore
except ImportError:
    try:                                              # legacy path
        from openpyxl.utils.cell import format_cell  # type: ignore
    except ImportError:                               # tiny fallback
        def format_cell(cell) -> str:  # type: ignore
            val = cell.value
            if val is None:
                return ""
            fmt = getattr(cell, "number_format", "")
            if not isinstance(val, (int, float)) or "." not in fmt:
                return str(val)
            decs = sum(ch == "0" for ch in fmt.split(".", 1)[1].split(";")[0])
            return f"{val:.{decs}f}"


def _clean_cell(cell) -> Any:
    """Return a value that preserves the cell’s displayed decimal places."""
    if cell.data_type != "n":                     # non-numeric
        return cell.value
    txt = format_cell(cell)                      # what Excel shows
    if "E" in txt or "e" in txt:                 # scientific notation – leave raw
        return cell.value
    if "." in txt:                               # round to shown decimals
        n_dec = len(txt.split(".", 1)[1])
        q = Decimal(cell.value).quantize(
            Decimal("1." + "0" * n_dec), rounding=ROUND_HALF_UP
        )
        return float(q)
    return cell.value


def _sanitize_headers(raw_headers: Sequence[Any]) -> List[str]:
    """
    Emulate pandas' behaviour:
      * NaN/None/''  →  'Unnamed: {idx}'
      * duplicate names → '.1', '.2', …
    """
    clean: List[str] = []
    counts = {}
    for idx, h in enumerate(raw_headers):
        name = f"Unnamed: {idx}" if pd.isna(h) or h == "" else str(h)
        if name in counts:
            counts[name] += 1
            name = f"{name}.{counts[name]}"
        else:
            counts[name] = 0
        clean.append(name)
    return clean


def read_excel_preserve_decimals(
    path: str,
    sheet_name: Any = 0,
    header: int | Sequence[int] | None = 0,
    **pd_kwargs,
) -> pd.DataFrame:
    """
    Load *path*/*sheet_name* into a DataFrame while

    - keeping the exact decimal precision the user sees in Excel,
    - reproducing pandas’ header processing rules.

    All extra keyword arguments are forwarded to ``pd.DataFrame`` after loading.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    rows: List[List[Any]] = [[_clean_cell(c) for c in row] for row in ws.iter_rows()]

    df = pd.DataFrame(rows, **pd_kwargs)

    # Apply header logic (only single-row headers are used in this project)
    if header is not None:
        df.columns = _sanitize_headers(df.iloc[header])
        df = df.drop(index=list(range(header + 1))).reset_index(drop=True)

    return df
